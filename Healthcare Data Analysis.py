# coding: utf-8

# In[3]:

#Project to analyze medicare data
#Importing required libraries

import requests, zipfile, os,openpyxl, sqlite3,glob, getpass,csv,pandas as pd, statistics
from openpyxl import load_workbook


# In[4]:

# Fetching data from medicare.gov using request package

url="https://data.medicare.gov/views/bg9k-emty/files/0a9879e0-3312-4719-a1db-39fd114890f1?content_type=application%2Fzip%3B%20charset%3Dbinary&filename=Hospital_Revised_Flatfiles.zip"
r=requests.get(url)                         #Checking different parameters of requests.get()
print (r.headers)
print(r.encoding)
print(r.headers['content-type'])
print(type(r.content))


# In[6]:

# Creating staging directory so that it will be helpful for debugging purpose
#Creating staging directory
staging_dir_name = "staging"
os.mkdir(staging_dir_name)
os.path.isdir(staging_dir_name)


# In[7]:

# Writing content to zip file
zip_file_name = os.path.join(staging_dir_name, "test.zip")
zf = open(zip_file_name,"wb")
zf.write(r.content)
zf.close()

#Extracting content of zip file to staging directory
z = zipfile.ZipFile(zip_file_name,'r')
z.extractall(staging_dir_name)
z.close()


# In[12]:

#Connecting to medicare database in sqlite 
#Fetching data using Python library Pandas

conn = sqlite3.connect("medicare_hospital_compare.db")
c1 = conn.cursor()

glob_dir = os.path.join(staging_dir_name,"*.csv")             #glob directory contains all csv files from staging
for file_name in glob.glob(glob_dir):                         #Looping through all files in staging 
    abspath = os.path.abspath(file_name)           
    df = pd.read_csv(abspath, header=None, encoding='cp1252')                      #Using pandas to read csv file
    
    
    #firstrow = df.iloc[0]                                                       #Fetching first row from dataframe
    split_ext = os.path.splitext(os.path.basename(file_name))[0]                 #File without extension 
    split_ext = split_ext.lower()                               #Applying transformation for inserting into sqlite
    split_ext = split_ext.replace(" ","_")
    split_ext = split_ext.replace("-","_")
    split_ext = split_ext.replace("%","pct")
    split_ext = split_ext.replace("/","_")
    if split_ext[0].isalpha() == False:                         #Appending t_ if not starting with alphabate 
         split_ext = "t_" + split_ext
    
    i = 0 
    table_name = split_ext                                    #Table name after processing
    sql_str = "drop table if exists %s" %(table_name)
    c1.execute(sql_str)
    
    new_file_name = table_name + '_new.csv'                     #For encoding purpose creating file with _new
    df.to_csv(new_file_name,encoding='utf-8', index = False)
    data = pd.read_csv(new_file_name, header=None, encoding='utf-8')
    firstrow = data.iloc[1]                                      #Fetching first row
    
    for val in firstrow:                                         #Applying transformation for columns
        val = val.lower()
        val = val.replace(" ","_")
        val = val.replace("-","_")
        val = val.replace("%","pct")
        val = val.replace("/","_")
        if val[0].isalpha() == False:                           #Appending c_ if not starting with alphabate 
             val = "c_" + val
        
        new_field = val
        field_type= 'Text'
        if i == 0:                                                            #If first column then create
            c1.execute('CREATE TABLE {tn} ({nf} {ft})'.format(tn=table_name,nf=new_field,ft=field_type))
            conn.commit()
        else:                                                                #Else alter table
            c1.execute("ALTER TABLE {tn} ADD COLUMN '{cn}' {ct}".format(tn=table_name,cn=new_field,ct=field_type))
        
        i+=1

    data1 = data.iloc[2:]                                               #Rows without header                                                             
    wildcards = ','.join(['?'] * len(data1.columns))
    insert_sql = 'INSERT INTO %s VALUES (%s)' %(table_name,wildcards)     #Inserting data to tables
    tup_data = [tuple(x) for x in data1.values]         
    c1.executemany(insert_sql, tup_data)                                
    conn.commit()                                                         #Commit

print("Table created and populated with values")    


# In[13]:

#First part was downloading data and inserting to tables. Table name was same as CSV file.
#Fetching data for hospital ranking.  
k_url = "http://kevincrook.com/utd/hospital_ranking_focus_states.xlsx"   
r = requests.get(k_url)                                                        
r.headersxf = open("hospital_ranking_fodcus_states.xlsx","wb")
xf = open("hospital_ranking_fodcus_states.xlsx","wb")
xf.write(r.content)                                    #Writing contents to excel file
xf.close()
print("Fetching data done") 


# In[14]:

#Check data from hospital_ranking_fodcus_states 

wb = openpyxl.load_workbook("hospital_ranking_fodcus_states.xlsx")
for sheet_name in wb.get_sheet_names():                               #Checking sheets
    print(sheet_name)

sheet = wb.get_sheet_by_name("Hospital National Ranking")             #checking data from ranking sheet
i = 1
while sheet.cell(row=i, column=1).value !=None:
    print(sheet.cell(row=i, column=1).value, "|", sheet.cell(row=i,column=2).value)
    i += 1
    
sheet2 = wb.get_sheet_by_name("Focus States")                         #Checking data from focus states
i = 1
while sheet2.cell(row=i, column=1).value !=None:                      #Looping through last row
    print(sheet2.cell(row=i, column=1).value, "|", sheet2.cell(row=i,column=2).value)
    i += 1
    
print("Checking sheets done")


# In[26]:

#Creating excel hospital ranking
#Hospital ranking excel file includes data from hospital general information table

#Creating hospital_ranking file and adding columns


wb2 = openpyxl.Workbook()                    
sheet_1 = wb2.create_sheet("Nationwide")                 
sheet_1.cell(row=1,column=1,value="Provider ID")                 #Creating columns in workbook
sheet_1.cell(row=1,column=2,value="Hospital Name")
sheet_1.cell(row=1,column=3,value="City")
sheet_1.cell(row=1,column=4,value="State")
sheet_1.cell(row=1,column=5,value="County")

wb2.remove_sheet(wb2.get_sheet_by_name('Sheet'))                #Removing extra sheet
wb2.save("hospital_ranking.xlsx")

# 'sheet' is from Excel hospital_ranking_focus_states

# Fetching records for top 100 provider id from database and inserting to excel
#excel will hold general information

i = 2
while i < 102:
    prov_id = sheet.cell(row=i, column=1).value
    sheet_1.cell(row=i, column= 1).value = prov_id  
    
    #Based on provider id fetching data from database and inserting into excel
    rows = c1.execute('select hospital_name,city,state,county_name from hospital_general_information where provider_id = ?',
                      (prov_id,))
    for row in rows:                                   #Storing data to excel
        sheet_1.cell(row=i, column= 2).value = row[0] 
        sheet_1.cell(row=i, column= 3).value = row[1]
        sheet_1.cell(row=i, column= 4).value = row[2]  
        sheet_1.cell(row=i, column= 5).value = row[3]
        wb2.save("hospital_ranking.xlsx")

    i += 1     #Next provider id
    
    
#Creating sheet for each state; purpose of this is to insert measures for each state
wb2 = openpyxl.load_workbook("hospital_ranking.xlsx")    
wb = openpyxl.load_workbook("hospital_ranking_fodcus_states.xlsx")
sheet2 = wb.get_sheet_by_name("Focus States")

i = 2
while i < 12:                                                 #Looping through all states
        hospital_state = sheet2.cell(row=i, column=1).value
        final_sheet = wb2.create_sheet(hospital_state)
        final_sheet.cell(row=1,column=1).value="Provider ID"      #Creating columns in excel
        final_sheet.cell(row=1,column=2).value="Hospital Name"
        final_sheet.cell(row=1,column=3).value="City"
        final_sheet.cell(row=1,column=4).value="State"
        final_sheet.cell(row=1,column=5).value="County"
        i += 1                                                  #Increamenting column                 

wb2.save("hospital_ranking.xlsx")
        
#Inserting data row-wise
    
tup_list = []
sheet = wb.get_sheet_by_name("Hospital National Ranking")
m = 2

while m < 12:                                                  #Looping through all states
    tup_list = []
    hospital_abs = sheet2.cell(row=m, column=2).value
    hospital_state = sheet2.cell(row=m, column=1).value
    
    #Fetching records based on state
    df = c1.execute("select provider_id,hospital_name,city,state,county_name from hospital_general_information where state = ?",(hospital_abs,))
    
    for rows in df: 
        con_list = list(rows)           #Converting tuple to rows
        inp_prov = con_list[0]          #provider_id
    
        i = 2
        while sheet.cell(row=i, column=1).value != None:
            rank = sheet.cell(row=i, column=2).value              #Taking corresponding rank from focus_state file
            
            if (sheet.cell(row=i, column=1).value == inp_prov):
                break
            i += 1
            
        con_list.append(rank)                 #Appending rank to the fetched record
        tup_list.append(con_list)             #Creating list of all fetched records
        
    tup_list.sort(key=lambda x: x[5])                #sorting records based on rank
    

    sheet3 = wb2.get_sheet_by_name(hospital_state)  #Sheet for each state
    print(sheet3)

    i=0
    j=2
    while i < 100:                                            #Transfering data from list to excel
        sheet3.cell(row=j, column=1).value = tup_list[i][0]
        sheet3.cell(row=j, column=2).value = tup_list[i][1]
        sheet3.cell(row=j, column=3).value = tup_list[i][2]
        sheet3.cell(row=j, column=4).value = tup_list[i][3]
        sheet3.cell(row=j, column=5).value = tup_list[i][4]
        
        
        i += 1                                                 # i is counter for traversing fetched records
        j += 1                                                 # j is for excel 
    
    wb2.save("hospital_ranking.xlsx")
    m = m + 1                         # m is counter for all states
    
wb2.save("hospital_ranking.xlsx")
print("Flow Done")


# In[63]:

# Creating file measures_statistics


wb3 = openpyxl.Workbook()
stat_nation = wb3.create_sheet("Nationwide")         #Sheet for nationwide data
stat_nation.cell(row=1,column=1).value="Measure ID"
stat_nation.cell(row=1,column=2).value="Measure Name"
stat_nation.cell(row=1,column=3).value="Minimum"
stat_nation.cell(row=1,column=4).value="Maximum"
stat_nation.cell(row=1,column=5).value="Average"
stat_nation.cell(row=1,column=6).value="Standard Deviation"    #Created columns in sheet

wb3.remove_sheet(wb3.get_sheet_by_name('Sheet'))
wb3.save("measures_statistics.xlsx")

#Creating sheet for each state
wb3 = openpyxl.load_workbook("measures_statistics.xlsx")
wb = openpyxl.load_workbook("hospital_ranking_fodcus_states.xlsx")
sheet2 = wb.get_sheet_by_name("Focus States")

#For each state calculate statistics in excel tab; name each tab with state name
i = 2
while i < 12:                                                      #Traversing through all states
        hospital_state = sheet2.cell(row=i, column=1).value        
        final_sheet = wb3.create_sheet(hospital_state)             #Creating sheet with name as state
        final_sheet.cell(row=1,column=1).value="Measure ID"
        final_sheet.cell(row=1,column=2).value="Measure Name"       
        final_sheet.cell(row=1,column=3).value="Minimum"
        final_sheet.cell(row=1,column=4).value="Maximum"
        final_sheet.cell(row=1,column=5).value="Average"                       #Creating columns
        final_sheet.cell(row=1,column=6).value="Standard Deviation"
        i += 1
        wb3.save("measures_statistics.xlsx")
        

wb3.save("measures_statistics.xlsx")
print("flow done")



# In[64]:

#Fetching distinct measure id from timely_and_effective_care_hospital table

conn = sqlite3.connect("medicare_hospital_compare.db")
measure_id = []            #Empty lists
measure_name = []
c1 = conn.cursor()                  
measure_sheet = wb3.get_sheet_by_name("Nationwide")     

#Fetching measure id and name
measure = c1.execute("select distinct measure_id,measure_name from timely_and_effective_care___hospital order by measure_id")
for row in measure:
            measure_id.append(row[0]) 
            measure_name.append(row[1])
#printing output        
print(measure_id)
print(measure_name)
print("Flow done")


# In[65]:

#Inserting data into measure_statistics.Nationwide sheet

#Defining function to check if score has numerical values
def is_number(s):
    try:
        complex(s) # for int, long, float and complex
    except ValueError:
        return False                                      #If not numerical then return false

    return True

#Connecting to data base
conn = sqlite3.connect("medicare_hospital_compare.db")
c1 = conn.cursor()
wb3 = load_workbook("measures_statistics.xlsx")
measure_sheet = wb3.get_sheet_by_name("Nationwide")


#Initializing variable
i=0
k = 2
temp = []
min_val = 0                   #setting output variables to 0
max_val = 0
std_dev = 0


#Looping through list of measure_id
while i < len(measure_id):
    temp = []                       #Initializing variables
    min_val = 0
    max_val = 0
    std_dev = 0
    get_id = measure_id[i]          #Current measure_id
        
    
    #Fetching data based on measure id
    scores = c1.execute("select score from timely_and_effective_care___hospital where measure_id =?",(get_id,))

    for score in scores:
        con_list = list(score)                      #Converting tuple to list
        
        check = is_number(con_list[0])              #Calling function to check numerical values 
        
        if check == True:                           #If value is numerical then only append 
            n = float(con_list[0])
            temp.append(n)
    
    
    if temp != []:                                  #If list not empty
        min_val = min(temp)                         #Calculate required values
        max_val = max(temp)
        avg = statistics.mean(temp)
        std_dev = statistics.stdev(temp)
        
        measure_sheet.cell(row=k, column= 1).value = measure_id[i]     #Inserting values into sheet
        measure_sheet.cell(row=k, column= 2).value = measure_name[i]
        measure_sheet.cell(row=k, column= 3).value = min_val
        measure_sheet.cell(row=k, column= 4).value = max_val
        measure_sheet.cell(row=k, column= 5).value = avg
        measure_sheet.cell(row=k, column= 6).value = std_dev
        wb3.save("measures_statistics.xlsx")                           #Saving excel        
        k = k + 1
            
    i += 1
print("Flow done")


# In[67]:

#Changing defination of execute to accept 3 parameters
def Execute(self, SQL, params=()):
    self.__connection.execute(SQL, params)
    self.__connection.commit()

wb3 = load_workbook("measures_statistics.xlsx")

j = 0                           #Initializing variables
i = 2
while i < 12:                                           #Looping through all states
        hospital_state = sheet2.cell(row=i, column=1).value
        hospital_abs = sheet2.cell(row=i, column=2).value
        state_sheet = wb3.get_sheet_by_name(hospital_state)
        j=0                                                #Initializing variables
        temp = []
        min_val = 0
        max_val = 0
        std_dev = 0
        k = 2 
        while j < len(measure_id):                         #Looping through all variables
            
            temp = []
            min_val = 0
            max_val = 0
            std_dev = 0
            avg = 0
            get_id = measure_id[j]                      #input to query
            
            #Fetching records based on measure_id and state
            scores = c1.execute("select score from timely_and_effective_care___hospital where measure_id =? and state=?",
                                (get_id,hospital_abs))
            for score in scores:
                con_list = list(score)                 #Converting tuple to list
                check = is_number(con_list[0])         #Check if the numerical values
                if check == True:
                    n = float(con_list[0])             #converting numerical string into float for calculation
                    temp.append(n)                     #Creating list of scores
            
            if temp != []:
                min_val = min(temp)                 #Calculating required values
                max_val = max(temp)
                avg = statistics.mean(temp)
                try:                                           
                    std_dev = statistics.stdev(temp)               #If any error assign std_dev = 0
                except :
                     std_dev = 0
                state_sheet.cell(row=k, column= 1).value = measure_id[j]          
                state_sheet.cell(row=k, column= 2).value = measure_name[j]
                state_sheet.cell(row=k, column= 3).value = min_val            #Storing values in excel
                state_sheet.cell(row=k, column= 4).value = max_val
                state_sheet.cell(row=k, column= 5).value = avg
                state_sheet.cell(row=k, column= 6).value = std_dev
                k += 1                              
            wb3.save("measures_statistics.xlsx")                              #Saving excel
            j += 1                                                 #Increamenting counter in measure_id list
        i += 1                                                     #Increamenting counter for states
        
print("flow_done")
wb3.save("measures_statistics.xlsx")
        




